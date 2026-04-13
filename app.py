import streamlit as st
import pandas as pd
import openpyxl
import io
import re
import zipfile
import numpy as np
from docxtpl import DocxTemplate

# =====================================================================
# 1. CẤU HÌNH HỆ THỐNG & TỪ ĐIỂN TKB (Dữ liệu của Trường Hoa Sen)
# =====================================================================
TU_DIEN = {
    ("Vân", "L"): "Cô Vân (Lý)",
    ("Vân", "CĐ Lý"): "Cô Vân (Lý)",
    ("Vân", "Đ"): "Cô Vân (Địa)",
    ("Vân", "A"): "Cô Thảo Vân (Anh)",
    ("Vân", "IELTS/A"): "Cô Thảo Vân (Anh)",
    ("Vân", "TNHN"): "Cô Vân (Lý)",
    ("Nhung", "V"): "Cô T.Nhung (Văn)",
    ("Nhung", "CĐ Văn"): "Cô T.Nhung (Văn)",
    ("Nhung", "A"): "Cô Nhung (Anh)",
    ("Nhung", "AVTH"): "Cô Nhung (Anh)",
    ("Nhung", "TNHN"): "Cô Nhung (TNHN)",
    ("Tâm", "V"): "Thầy Tâm (Văn)",
    ("Tâm", "CĐ Văn"): "Thầy Tâm (Văn)",
    ("Tâm", "AVTH"): "Cô Tâm (Anh)",
    ("Ngọc", "L"): "Cô Ngọc (Lý)",
    ("Ngọc", "CĐ Lý"): "Cô Ngọc (Lý)",
    ("Ngọc", "KTPL"): "Thầy Ngọc (KTPL)",
    ("Ngọc", "CĐ KTPL"): "Thầy Ngọc (KTPL)",
    ("Ngọc", "CN"): "Cô Ngọc (Công Nghệ)",
    ("Phương", "V"): "Cô Phương (Văn)",
    ("Phương", "Su"): "Cô Phương (Sử)",
    ("Phương", "CĐ Sử"): "Cô Phương (Sử)",
    ("Anh", "KTPL"): "Cô Lan Anh (GDCD/KTPL)",
    ("Anh", "CĐ KTPL"): "Cô Lan Anh (GDCD/KTPL)",
    ("Anh", "GDCD"): "Cô Lan Anh (GDCD/KTPL)",
    ("Nghĩa", "T"): "Thầy Nghĩa (Toán)",
    ("Nghĩa", "CĐ Toán"): "Thầy Nghĩa (Toán)",
    ("nghĩa", "CĐ Toán"): "Thầy Nghĩa (Toán)",
    ("Bình", "V"): "Thầy/Cô Bình (Văn)",
    ("Bình", "CĐ Văn"): "Thầy/Cô Bình (Văn)",
    ("Bảo", "Su"): "Thầy Bảo (Sử/GDĐP)",
    ("Bảo", "SĐ"): "Thầy Bảo (Sử/GDĐP)",
    ("Chi", "H"): "Cô Chi (Hóa)",
    ("Chi", "TNHN"): "Cô Chi (Hóa)",
    ("Diệp", "V"): "Cô Diệp (Văn)",
    ("Diệp", "CĐ Văn"): "Cô Diệp (Văn)",
    ("Diệp", "Su"): "Cô Diệp (Sử)",
    ("Diệp", "GDĐP"): "Cô Diệp (Sử)",
    ("Xuân", "GDCD"): "Cô Xuân (GDCD/KTPL)",
    ("Xuân", "KTPL"): "Cô Xuân (GDCD/KTPL)",
    ("Xuân", "CĐ KTPL"): "Cô Xuân (GDCD/KTPL)",
    ("Vinh", "GDCD"): "Cô Vinh (GDCD/KTPL)",
    ("Vinh", "KTPL"): "Cô Vinh (GDCD/KTPL)",
    ("Vinh", "CĐ KTPL"): "Cô Vinh (GDCD/KTPL)",
}

def get_standard_name(ten_ky_hieu, mon_hoc):
    if mon_hoc == 'Chủ nhiệm':
        return f"{ten_ky_hieu} (GVCN)"
    if (ten_ky_hieu, mon_hoc) in TU_DIEN:
        return TU_DIEN[(ten_ky_hieu, mon_hoc)]
    return ten_ky_hieu

def phan_loai_khoi(ten_lop):
    ten_lop = str(ten_lop).strip()
    if ten_lop.startswith(('10', '11', '12')):
        return 'THPT'
    elif ten_lop.startswith(('6', '7', '8', '9')):
        return 'THCS'
    return 'Khác'

# =====================================================================
# 2. LOGIC XỬ LÝ DỮ LIỆU TKB
# =====================================================================
def process_tkb_data(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_cell_value
        if ws.max_row > 66:
            ws.delete_rows(67, ws.max_row - 66)
            
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    df = pd.read_excel(virtual_workbook) 

    classes, gvcn = {}, {}
    class_row = df.iloc[3]
    gvcn_row = df.iloc[4]
    
    for col_idx, val in enumerate(class_row):
        if pd.notna(val) and isinstance(val, str) and re.match(r'^(1[0-2]|[6-9])', val.strip()): 
            class_name = val.strip()
            classes[col_idx] = class_name
            gv_val = str(gvcn_row.iloc[col_idx]).strip()
            gvcn[class_name] = gv_val.split('-')[0].strip() if '-' in gv_val else gv_val

    records = []
    for row_idx in range(5, len(df)):
        row = df.iloc[row_idx]
        for col_idx, class_name in classes.items():
            cell_val = str(row.iloc[col_idx]).strip()
            if cell_val in ['nan', '', 'CHÀO CỜ', 'SINH HOẠT ĐẦU GIỜ', 'THỂ DỤC THỂ THAO']: continue
            
            khoi = phan_loai_khoi(class_name)
            if cell_val.lower() == 'chủ nhiệm':
                ten_goc = gvcn.get(class_name, 'Unknown')
                records.append({'Khối': khoi, 'Giáo viên': get_standard_name(ten_goc, 'Chủ nhiệm'), 'Lớp': class_name, 'Môn': 'Chủ nhiệm'})
            elif '-' in cell_val:
                parts = cell_val.split('-')
                mon_hoc, ten_goc = "-".join(parts[:-1]).strip(), parts[-1].strip()
                records.append({'Khối': khoi, 'Giáo viên': get_standard_name(ten_goc, mon_hoc), 'Lớp': class_name, 'Môn': mon_hoc})
            else:
                records.append({'Khối': khoi, 'Giáo viên': 'Chung', 'Lớp': class_name, 'Môn': cell_val})

    df_res = pd.DataFrame(records)
    return df_res.groupby(['Khối', 'Giáo viên', 'Lớp', 'Môn']).size().reset_index(name='Số tiết') if not df_res.empty else None

# =====================================================================
# 3. GIAO DIỆN STREAMLIT & ĐIỀU HƯỚNG
# =====================================================================
st.set_page_config(page_title="HSE Admin Tool", page_icon="🏫", layout="wide")

if 'page' not in st.session_state:
    st.session_state.page = "Main"

def go_home(): st.session_state.page = "Main"

# --- PAGE: MAIN MENU ---
if st.session_state.page == "Main":
    st.title("🏫 Hệ thống Quản trị Trường Hoa Sen")
    st.write("Chào mừng Admin. Vui lòng chọn ứng dụng cần xử lý:")
    
    c1, c2 = st.columns(2)
    with c1:
        st.info("### 📊 Thống kê TKB")
        st.write("Phân tích dữ liệu Thời khóa biểu, tính số tiết theo giáo viên.")
        if st.button("Mở Thống kê TKB"): 
            st.session_state.page = "TKB"
            st.rerun()
    with c2:
        st.success("### 🎓 Báo Cáo Điểm")
        st.write("Trộn điểm từ Excel vào file Word hàng loạt và xuất ZIP.")
        if st.button("Mở Báo Cáo Điểm"): 
            st.session_state.page = "Report"
            st.rerun()

# --- PAGE: THỐNG KÊ TKB ---
elif st.session_state.page == "TKB":
    st.title("📊 Thống kê Thời khóa biểu")
    if st.button("⬅️ Quay lại Menu"): go_home(); st.rerun()
    
    file_tkb = st.file_uploader("Tải file TKB (.xlsx)", type=["xlsx"])
    if file_tkb:
        if st.button("Phân tích dữ liệu"):
            df_tkb = process_tkb_data(file_tkb)
            if df_tkb is not None:
                st.session_state.df_tkb = df_tkb
                st.success("Đã xử lý xong!")
            else: st.error("Không có dữ liệu.")

    if 'df_tkb' in st.session_state:
        df = st.session_state.df_tkb
        col_f1, col_f2 = st.columns(2)
        gv_filter = col_f1.multiselect("Lọc Giáo viên", sorted(df['Giáo viên'].unique()))
        khoi_filter = col_f2.multiselect("Lọc Khối", sorted(df['Khối'].unique()))
        
        df_view = df.copy()
        if gv_filter: df_view = df_view[df_view['Giáo viên'].isin(gv_filter)]
        if khoi_filter: df_view = df_view[df_view['Khối'].isin(khoi_filter)]
        
        st.dataframe(df_view, use_container_width=True)
        st.metric("Tổng số tiết", df_view['Số tiết'].sum())
        csv = df_view.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
        st.download_button("📥 Tải kết quả CSV", csv, "ThongKeTKB.csv", "text/csv")

# --- PAGE: BÁO CÁO ĐIỂM ---
elif st.session_state.page == "Report":
    st.title("🎓 Báo Cáo Điểm (Mail Merge)")
    if st.button("⬅️ Quay lại Menu"): go_home(); st.rerun()
    
    col_u1, col_u2 = st.columns(2)
    word_file = col_u1.file_uploader("1. Chọn mẫu Word (template.docx)", type=["docx"])
    excel_file = col_u2.file_uploader("2. Chọn file điểm (Excel)", type=["xlsx"])
    
    if word_file and excel_file:
        df_score = pd.read_excel(excel_file).fillna("")
        df_score.columns = [c.replace(' ', '_').replace('\n', '_') for c in df_score.columns]
        
        if st.button("🚀 Bắt đầu trộn và tạo file ZIP"):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for _, row in df_score.iterrows():
                    if not row['Họ_và_Tên']: continue
                    doc = DocxTemplate(word_file)
                    # Định dạng số 2 chữ số thập phân cho các cột điểm
                    context = {}
                    for k, v in row.to_dict().items():
                        try:
                            val = float(v)
                            context[k] = f"{val:.2f}"
                        except:
                            context[k] = v
                    
                    doc.render(context)
                    doc_io = io.BytesIO()
                    doc.save(doc_io)
                    file_name = f"{row['Họ_và_Tên']} - {row.get('Lớp', 'Lớp')}.docx"
                    zf.writestr(file_name, doc_io.getvalue())
            
            st.balloons()
            st.success("Đã trộn xong!")
            st.download_button("📥 Tải về file ZIP kết quả", zip_buffer.getvalue(), "PhieuDiem_HSE.zip")
